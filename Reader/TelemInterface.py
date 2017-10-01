import sys
import serial
from io import StringIO
import csv
from openpyxl import Workbook
import datetime
from dateutil.parser import parse



from PyQt5.QtGui import QGuiApplication
from PyQt5.QtQml import QQmlApplicationEngine
from PyQt5.QtCore import QObject, pyqtSignal, pyqtSlot, QThread



global ser

# Create an instance of the application
app = QGuiApplication(sys.argv)
#define and open the serial port
ser=serial.Serial('COM6')
# Create QML engine
engine = QQmlApplicationEngine()


class Dash(QObject):
    def __init__(self):
        QObject.__init__(self)
   #create the signals that change the GUI     
    ampHourvalue = pyqtSignal(float, arguments=['amphour'])
    auxVoltage = pyqtSignal(float, arguments=['auxvolt'])
    mainVoltage = pyqtSignal(float, arguments=['mainvolt'])
    arrayCurrent = pyqtSignal(float, arguments=['arraycurrent'])
    motorCurrent = pyqtSignal(float, arguments=['motorcur'])

    
    @pyqtSlot(int)
    def reset(self, value):
        
        value=100  #put amp hour capacity here
        self.ampHourvalue.emit(value)


# Create a Dash object
dashboard = Dash()


class ThreadClass(QThread):
    # Create the signal
    ampHourvalue = pyqtSignal(float, arguments=['amphour'])
    auxVoltage = pyqtSignal(float, arguments=['auxvolt'])
    mainVoltage = pyqtSignal(float, arguments=['mainvolt'])
    arrayCurrent = pyqtSignal(float, arguments=['arraycurrent'])
    motorCurrent = pyqtSignal(float, arguments=['motorcur'])
    

    def __init__(self, parent=None):
        super(ThreadClass, self).__init__(parent)
        #connects  the signals from the thread to the signals in the thread that is running the GUI
        
        self.auxVoltage.connect(dashboard.auxVoltage)
        self.mainVoltage.connect(dashboard.mainVoltage)
        self.motorCurrent.connect(dashboard.motorCurrent)
        self.arrayCurrent.connect(dashboard.arrayCurrent)
        self.ampHourvalue.connect(dashboard.ampHourvalue)
        


    def run(self):
        
        #This implementation of amp hours needs work
        ampHour=100  #start with a full tank of electrons.
        
        try:
            wb=load_workbook("History.xls") #attemps to open the history excel file
        except:
            wb=Workbook() #creates and empty excel workbook if histortory is not found
    
        WorkSheetName=datetime.date.today() #get todays date
        ws = wb.create_sheet("%s" %WorkSheetName) #create worksheet with the date as tittle
        
        firstrun=True  # logical flag to prevent errors on first run
        while True:
            
            

            data=ser.readline(120) #read the stream
            data=data.decode() #convert stream from bytes to characters
            data=StringIO(data)#convert a stream of characters into string
            dataset=csv.reader(data, delimiter= ',') #read the CSV string into individual array
            dataset=list(dataset) #convert the array to list
            
            #write to excel log
            ws.append(dataset[0])
            wb.save("History.xls")
            
            #extract individual data points
            timestamp=(dataset[0][0])
            aux=float(dataset[0][1])
            mainvolt=float(dataset[0][2])
            motorcurrent=float(dataset[0][3])
            arraycurrent=float(dataset[0][4])
            
            # Emit the signals
            self.auxVoltage.emit(aux)
            self.mainVoltage.emit(mainvolt)
            self.arrayCurrent.emit(arraycurrent) 
            self.motorCurrent.emit(motorcurrent)
            
            
######## Calculate amp hours######
#room for growth.  Should include Puekert equation
            
            #don't call the subtraction on the first run, old timestamp does not exist
            if firstrun:
                firstrun=False
                timestamp=parse(timestamp)
            
            #calc time between last two records    
            else:
               timestamp=parse(timestamp)
               timedelta=timestamp-oldtimestamp               
               #calculate the number of amphours used on this interval
               currentused= (motorcurrent-arraycurrent)*(float(timedelta.seconds)/3600)

               #subtract the number of amphours used from the remaining amount
               ampHour= ampHour - currentused
               #send to the guage
               self.ampHourvalue.emit(ampHour)
               print(ampHour)
            
            
            oldtimestamp=timestamp
            
            
            

            
            
             
            
        pass



#start the thread
threadclass=ThreadClass()
threadclass.start()
# And register it in the context of QML
engine.rootContext().setContextProperty("dashboard", dashboard)
# Load the qml file into the engine
engine.load("take2.qml")
engine.quit.connect(app.quit)
sys.exit(app.exec_())
        
      

       
    
        
